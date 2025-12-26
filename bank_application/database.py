import sqlite3
import json
import hashlib
import secrets
import os
import base64
from datetime import datetime
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

class CompleteDatabase:
    """Complete database with all features"""

    def __init__(self, master_password=None):
        self.master_password = master_password
        # Use absolute path relative to the script location
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.db_path = os.path.join(base_dir, 'complete_bank_manager.db')
        self._init_database()

    def _init_database(self):
        """Initialize database with all fields"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Create settings table for password hash
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        ''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS bank_cards (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_name TEXT NOT NULL,
                branch_name TEXT NOT NULL,
                ifsc_code TEXT NOT NULL,
                account_number TEXT NOT NULL,
                atm_number TEXT NOT NULL,
                pin TEXT NOT NULL,
                validity_start TEXT NOT NULL,
                validity_end TEXT NOT NULL,
                cvv TEXT NOT NULL,
                card_type TEXT NOT NULL,
                card_network TEXT NOT NULL,
                family_member TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        ''')

        conn.commit()
        conn.close()

    def is_setup_complete(self) -> bool:
        """Check if master password is set"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT value FROM settings WHERE key = "password_hash"')
        result = cursor.fetchone()
        conn.close()
        return result is not None

    def set_master_password(self, password: str) -> None:
        """Set master password"""
        salt = secrets.token_hex(16)
        # Use PBKDF2 with SHA256, 100,000 iterations
        password_hash = hashlib.pbkdf2_hmac(
            'sha256',
            password.encode('utf-8'),
            salt.encode('utf-8'),
            100000
        ).hex()
        stored_value = f"{salt}:{password_hash}"

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)',
                      ("password_hash", stored_value))
        conn.commit()
        conn.close()
        self.master_password = password

    def check_master_password(self, password: str) -> bool:
        """Check if password is correct"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT value FROM settings WHERE key = "password_hash"')
        result = cursor.fetchone()
        conn.close()

        if not result:
            return False

        salt, stored_hash = result[0].split(':')
        calculated_hash = hashlib.pbkdf2_hmac(
            'sha256',
            password.encode('utf-8'),
            salt.encode('utf-8'),
            100000
        ).hex()

        if calculated_hash == stored_hash:
            self.master_password = password
            return True
        return False

    def _get_encryption_key(self) -> bytes:
        """Derive encryption key from master password"""
        if not self.master_password:
            raise ValueError("Master password not set or logged in")

        # We need a consistent salt for the encryption key.
        # Ideally stored in DB, but for now we'll use the password hash's salt if available,
        # or a fixed system salt if we want to separate auth from encryption (better).
        # However, to avoid schema changes for a new 'encryption_salt', we'll read the auth salt.

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT value FROM settings WHERE key = "password_hash"')
        result = cursor.fetchone()
        conn.close()

        if not result:
             raise ValueError("Setup not complete")

        salt_hex = result[0].split(':')[0]
        salt = salt_hex.encode('utf-8') # Using the hex string as bytes for KDF salt

        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=salt,
            iterations=100000,
        )
        key = base64.urlsafe_b64encode(kdf.derive(self.master_password.encode()))
        return key

    def encrypt_data(self, data: str) -> str:
        """Encrypt string data"""
        if not data:
            return ""
        f = Fernet(self._get_encryption_key())
        return f.encrypt(data.encode()).decode()

    def decrypt_data(self, encrypted_data: str) -> str:
        """Decrypt string data"""
        if not encrypted_data:
            return ""
        try:
            f = Fernet(self._get_encryption_key())
            return f.decrypt(encrypted_data.encode()).decode()
        except Exception:
            # Fallback for legacy plain text data if migration happened
            return encrypted_data

    def mask_card_number(self, card_number: str) -> str:
        """Mask card number"""
        if len(card_number) >= 16:
            return f"{card_number[:4]} **** **** {card_number[-4:]}"
        return "**** **** **** ****"

    def mask_cvv(self, cvv: str) -> str:
        """Mask CVV"""
        if len(cvv) == 3:
            return "***"
        return "**"

    def add_card(self, bank_name, branch_name, ifsc_code, account_number, atm_number, pin,
                 validity_start, validity_end, cvv, card_type, card_network, family_member) -> int:
        """Add new bank card"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        now = datetime.now().isoformat()

        # Encrypt sensitive fields
        enc_account = self.encrypt_data(account_number)
        enc_atm = self.encrypt_data(atm_number)
        enc_pin = self.encrypt_data(pin)
        enc_cvv = self.encrypt_data(cvv)

        cursor.execute('''
            INSERT INTO bank_cards
            (bank_name, branch_name, ifsc_code, account_number, atm_number, pin,
             validity_start, validity_end, cvv, card_type, card_network,
             family_member, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            bank_name, branch_name, ifsc_code, enc_account, enc_atm, enc_pin,
            validity_start, validity_end, enc_cvv, card_type, card_network,
            family_member, now, now
        ))

        card_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return card_id

    def get_all_cards(self) -> list[dict]:
        """Get all cards with masked sensitive data"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute('''
            SELECT id, bank_name, branch_name, ifsc_code, account_number, atm_number,
                   pin, validity_start, validity_end, cvv, card_type, card_network,
                   family_member, created_at, updated_at
            FROM bank_cards
            ORDER BY created_at DESC
        ''')

        cards = []
        for row in cursor.fetchall():
            # Decrypt fields first
            dec_account = self.decrypt_data(row[4])
            dec_atm = self.decrypt_data(row[5])
            dec_pin = self.decrypt_data(row[6])
            dec_cvv = self.decrypt_data(row[9])

            card_data = {
                'id': row[0],
                'bank_name': row[1],
                'branch_name': row[2],
                'ifsc_code': row[3],
                'account_number': self.mask_card_number(dec_account),
                'atm_number': self.mask_card_number(dec_atm),
                'pin': self.mask_cvv(dec_pin),
                'validity_start': row[7],
                'validity_end': row[8],
                'cvv': self.mask_cvv(dec_cvv),
                'card_type': row[10],
                'card_network': row[11],
                'family_member': row[12],
                'created_at': row[13],
                'updated_at': row[14]
            }
            cards.append(card_data)

        conn.close()
        return cards

    def get_card_by_id(self, card_id: int) -> dict | None:
        """Get card by ID for editing"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute('''
            SELECT id, bank_name, branch_name, ifsc_code, account_number, atm_number,
                   pin, validity_start, validity_end, cvv, card_type, card_network,
                   family_member, created_at, updated_at
            FROM bank_cards WHERE id = ?
        ''', (card_id,))

        row = cursor.fetchone()
        conn.close()

        if row:
            return {
                'id': row[0],
                'bank_name': row[1],
                'branch_name': row[2],
                'ifsc_code': row[3],
                'account_number': self.decrypt_data(row[4]),
                'atm_number': self.decrypt_data(row[5]),
                'pin': self.decrypt_data(row[6]),
                'validity_start': row[7],
                'validity_end': row[8],
                'cvv': self.decrypt_data(row[9]),
                'card_type': row[10],
                'card_network': row[11],
                'family_member': row[12],
                'created_at': row[13],
                'updated_at': row[14]
            }
        return None

    def update_card(self, card_id, bank_name, branch_name, ifsc_code, account_number,
                   atm_number, pin, validity_start, validity_end, cvv, card_type,
                   card_network, family_member) -> None:
        """Update existing card"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        now = datetime.now().isoformat()

        # Encrypt sensitive fields
        enc_account = self.encrypt_data(account_number)
        enc_atm = self.encrypt_data(atm_number)
        enc_pin = self.encrypt_data(pin)
        enc_cvv = self.encrypt_data(cvv)

        cursor.execute('''
            UPDATE bank_cards
            SET bank_name=?, branch_name=?, ifsc_code=?, account_number=?, atm_number=?,
                pin=?, validity_start=?, validity_end=?, cvv=?, card_type=?,
                card_network=?, family_member=?, updated_at=?
            WHERE id=?
        ''', (
            bank_name, branch_name, ifsc_code, enc_account, enc_atm, enc_pin,
            validity_start, validity_end, enc_cvv, card_type, card_network,
            family_member, now, card_id
        ))

        conn.commit()
        conn.close()

    def delete_card(self, card_id: int) -> None:
        """Delete card by ID"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute('DELETE FROM bank_cards WHERE id = ?', (card_id,))
        conn.commit()
        conn.close()

    def get_all_cards_unmasked(self) -> list[dict]:
        """Get all cards with unmasked sensitive data for export"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute('''
            SELECT id, bank_name, branch_name, ifsc_code, account_number, atm_number,
                   pin, validity_start, validity_end, cvv, card_type, card_network,
                   family_member, created_at, updated_at
            FROM bank_cards
            ORDER BY created_at DESC
        ''')

        cards = []
        for row in cursor.fetchall():
            # Decrypt fields
            dec_account = self.decrypt_data(row[4])
            dec_atm = self.decrypt_data(row[5])
            dec_pin = self.decrypt_data(row[6])
            dec_cvv = self.decrypt_data(row[9])

            card_data = {
                'id': row[0],
                'bank_name': row[1],
                'branch_name': row[2],
                'ifsc_code': row[3],
                'account_number': dec_account,  # Unmasked
                'atm_number': dec_atm,      # Unmasked
                'pin': dec_pin,             # Unmasked
                'validity_start': row[7],
                'validity_end': row[8],
                'cvv': dec_cvv,             # Unmasked
                'card_type': row[10],
                'card_network': row[11],
                'family_member': row[12],
                'created_at': row[13],
                'updated_at': row[14]
            }
            cards.append(card_data)

        conn.close()
        return cards

    def export_to_excel(self, filename: str) -> int:
        """Export cards to Excel file"""
        # Note: This requires openpyxl installed
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        cards = self.get_all_cards_unmasked()

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Cards"

        # Define headers
        headers = [
            "ID", "Bank Name", "Branch Name", "IFSC Code", "Account Number",
            "ATM Number", "PIN", "Valid From", "Valid Until", "CVV",
            "Card Type", "Card Network", "Family Member", "Created At", "Updated At"
        ]

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add data rows
        for row, card in enumerate(cards, 2):
            ws.cell(row=row, column=1, value=card['id'])
            ws.cell(row=row, column=2, value=card['bank_name'])
            ws.cell(row=row, column=3, value=card['branch_name'])
            ws.cell(row=row, column=4, value=card['ifsc_code'])
            ws.cell(row=row, column=5, value=card['account_number'])
            ws.cell(row=row, column=6, value=card['atm_number'])
            ws.cell(row=row, column=7, value=card['pin'])
            ws.cell(row=row, column=8, value=card['validity_start'])
            ws.cell(row=row, column=9, value=card['validity_end'])
            ws.cell(row=row, column=10, value=card['cvv'])
            ws.cell(row=row, column=11, value=card['card_type'])
            ws.cell(row=row, column=12, value=card['card_network'])
            ws.cell(row=row, column=13, value=card['family_member'])
            ws.cell(row=row, column=14, value=card['created_at'])
            ws.cell(row=row, column=15, value=card['updated_at'])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws['A1'] = "Bank Cards Export Summary"
        summary_ws['A1'].font = Font(bold=True, size=14)

        summary_ws['A3'] = "Export Date:"
        summary_ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        summary_ws['A4'] = "Total Cards:"
        summary_ws['B4'] = len(cards)

        summary_ws['A6'] = "Card Types:"
        card_types = {}
        for card in cards:
            card_type = card['card_type']
            card_types[card_type] = card_types.get(card_type, 0) + 1

        row = 7
        for card_type, count in card_types.items():
            summary_ws[f'A{row}'] = f"{card_type}:"
            summary_ws[f'B{row}'] = count
            row += 1

        # Save the workbook
        wb.save(filename)
        return len(cards)

    def import_from_excel(self, filename: str) -> tuple[int, int]:
        """Import cards from Excel file"""
        from openpyxl import load_workbook
        try:
            wb = load_workbook(filename, data_only=True)
            ws = wb.active

            imported_count = 0
            skipped_count = 0

            # Skip header row, start from row 2
            for row in range(2, ws.max_row + 1):
                try:
                    # Read data from Excel
                    bank_name = str(ws.cell(row=row, column=2).value or "").strip()
                    branch_name = str(ws.cell(row=row, column=3).value or "").strip()
                    ifsc_code = str(ws.cell(row=row, column=4).value or "").strip()
                    account_number = str(ws.cell(row=row, column=5).value or "").strip()
                    atm_number = str(ws.cell(row=row, column=6).value or "").strip()
                    pin = str(ws.cell(row=row, column=7).value or "").strip()
                    validity_start = str(ws.cell(row=row, column=8).value or "").strip()
                    validity_end = str(ws.cell(row=row, column=9).value or "").strip()
                    cvv = str(ws.cell(row=row, column=10).value or "").strip()
                    card_type = str(ws.cell(row=row, column=11).value or "Debit").strip()
                    card_network = str(ws.cell(row=row, column=12).value or "RuPay").strip()
                    family_member = str(ws.cell(row=row, column=13).value or "").strip()

                    # Validate required fields
                    if not all([bank_name, branch_name, ifsc_code, account_number, atm_number, pin, cvv, validity_start, validity_end, family_member]):
                        skipped_count += 1
                        continue

                    # Clean account and ATM numbers (remove spaces)
                    account_number = ''.join(filter(str.isdigit, account_number))
                    atm_number = ''.join(filter(str.isdigit, atm_number))

                    # Validate ATM number length
                    if len(atm_number) != 16:
                        skipped_count += 1
                        continue

                    # Validate date format
                    try:
                        datetime.strptime(validity_start, "%Y-%m-%d")
                        datetime.strptime(validity_end, "%Y-%m-%d")
                    except ValueError:
                        skipped_count += 1
                        continue

                    # Add card to database
                    self.add_card(
                        bank_name, branch_name, ifsc_code, account_number, atm_number, pin,
                        validity_start, validity_end, cvv, card_type, card_network, family_member
                    )
                    imported_count += 1

                except Exception as e:
                    skipped_count += 1
                    continue

            return imported_count, skipped_count

        except Exception as e:
            raise Exception(f"Failed to import Excel file: {e}")

    def export_to_json(self, filename: str) -> int:
        """Export cards to JSON file (legacy support)"""
        cards = self.get_all_cards_unmasked()

        export_data = {
            'export_date': datetime.now().isoformat(),
            'version': '1.0.0',
            'total_cards': len(cards),
            'cards': cards
        }

        with open(filename, 'w') as f:
            json.dump(export_data, f, indent=2)

        return len(cards)
